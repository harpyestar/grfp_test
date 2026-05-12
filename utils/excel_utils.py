"""
Excel 文件生成与验证工具
用于生成签约状态导入 Excel 文件，以及验证导出的酒店名单 Excel
"""

import tempfile
from pathlib import Path
import openpyxl
import allure
from utils.config import config
from utils.logger import get_logger

logger = get_logger(__name__)

# 默认表头
DEFAULT_HEADER = ["HotelId", "Contract Status", "Remark"]


def generate_signing_status_excel(
    header: list = None,
    data_rows: list = None,
    file_name: str = "RFP_import_signing_status_file.xlsx",
) -> str:
    """
    生成签约状态导入 Excel 文件

    Args:
        header: 表头列表，默认 ["HotelId", "Contract Status", "Remark"]
        data_rows: 数据行列表，每行是一个列表，如 [["10086", "New", ""]]
        file_name: 输出文件名，默认 "RFP_import_signing_status_file.xlsx"

    Returns:
        str: Excel 文件的绝对路径
    """
    if header is None:
        header = DEFAULT_HEADER
    if data_rows is None:
        data_rows = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)

    temp_dir: Path = config.DATA_DIR / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    file_path = temp_dir / file_name

    if file_path.exists():
        try:
            file_path.unlink()
        except PermissionError:
            logger.warning(f"旧文件被占用，无法删除: {file_path}")

    wb.save(str(file_path))
    logger.info(f"签约状态导入文件已生成: {file_path}")
    return str(file_path)


def _read_excel_data(file_path: str) -> tuple:
    """
    读取 Excel 文件并返回表头和数据行（支持 .xlsx）

    Args:
        file_path: Excel 文件路径

    Returns:
        tuple: (headers: list, rows: list[list])
    """
    if not file_path.lower().endswith(".xlsx"):
        raise ValueError(f"不支持的 Excel 文件格式（仅支持 .xlsx）: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_values = [str(v).strip() if v is not None else "" for v in row]
        rows.append(row_values)
    wb.close()
    return headers, rows


async def verify_exported_hotel_excel(download, db_data: list, sheet_name: str) -> None:
    """
    验证导出的酒店名单 Excel 文件

    检查点:
    1. Excel 中存在酒店 ID 字段
    2. Excel 中每一行的酒店 ID 和名称与数据库查询结果一致

    Args:
        download: Playwright Download 对象
        db_data: 数据库查询结果列表 [{hotel_id, hotel_name}, ...]
        sheet_name: 表名（用于日志和 allure 报告）
    """
    # 保存下载文件到临时路径
    temp_dir = Path(tempfile.gettempdir()) / "rfp_export"
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / download.suggested_filename
    await download.save_as(str(temp_path))

    # 读取 Excel
    headers, excel_rows = _read_excel_data(str(temp_path))

    # 查找酒店 ID 列和酒店名列索引
    hotel_id_col = None
    hotel_name_col = None
    id_keywords = ["房仓酒店id", "hotelid", "酒店id", "酒店ID", "hotel_id", "hotel id"]
    name_keywords = ["酒店名称", "酒店名", "chn_name", "hotel_name", "name_en_us", "hotel name"]

    for col_idx, header in enumerate(headers):
        header_lower = str(header).lower().replace(" ", "").replace("_", "")
        for kw in id_keywords:
            if kw.lower().replace(" ", "").replace("_", "") in header_lower:
                hotel_id_col = col_idx
                break
        for kw in name_keywords:
            if kw.lower().replace(" ", "").replace("_", "") in header_lower:
                hotel_name_col = col_idx
                break

    # ===== 断言 1: 存在酒店 ID 列 =====
    assert hotel_id_col is not None, (
        f"[{sheet_name}] Excel 中未找到酒店 ID 字段！\n"
        f"表头: {headers}"
    )
    allure.attach(
        f"Excel 表头: {headers}\n酒店 ID 列索引: {hotel_id_col}\n"
        f"酒店名称列索引: {hotel_name_col}",
        f"{sheet_name} - Excel 结构",
        allure.attachment_type.TEXT,
    )

    # ===== 断言 2: 数据行数一致 =====
    assert len(excel_rows) == len(db_data), (
        f"[{sheet_name}] 数据行数不一致！\n"
        f"Excel 行数: {len(excel_rows)}, 数据库行数: {len(db_data)}"
    )

    # ===== 断言 3: 以 hotel_id 为 key 逐条匹配（无视顺序） =====
    db_dict = {row["hotel_id"]: row for row in db_data}
    mismatches = []
    matched_db_ids = set()

    for i, excel_row in enumerate(excel_rows):
        excel_id = excel_row[hotel_id_col].replace(".0", "")
        db_row = db_dict.get(excel_id)

        if db_row is None:
            mismatches.append(f"行 {i + 2}: Excel(ID={excel_id}) 在数据库中未找到")
            continue

        matched_db_ids.add(excel_id)
        db_id = db_row["hotel_id"]

        if excel_id != db_id:
            mismatches.append(f"行 {i + 2}: Excel(ID={excel_id}) ≠ DB(ID={db_id})")

        # 如果找到了酒店名列，对比名称
        if hotel_name_col is not None:
            excel_name = excel_row[hotel_name_col]
            db_name = db_row["hotel_name"]
            if excel_name != db_name:
                mismatches.append(
                    f"行 {i + 2}: Excel(名称={excel_name}) ≠ DB(名称={db_name})"
                )

    # ===== 断言 4: DB 数据在 Excel 中都有体现 =====
    missing_in_excel = set(db_dict.keys()) - matched_db_ids
    if missing_in_excel:
        mismatches.append(
            f"数据库中存在但 Excel 中缺失的酒店 ID ({len(missing_in_excel)} 个): "
            f"{', '.join(sorted(missing_in_excel))}"
        )

    assert len(mismatches) == 0, (
        f"[{sheet_name}] 数据不匹配！共 {len(mismatches)} 处:\n" + "\n".join(mismatches)
    )

    allure.attach(
        f"[{sheet_name}] 验证通过\n"
        f"- 共 {len(excel_rows)} 行数据\n"
        f"- 酒店 ID 字段: {'✓ 存在' if hotel_id_col is not None else '✗ 缺失'}\n"
        f"- 酒店名称字段: {'✓ 存在' if hotel_name_col is not None else '✗ 缺失'}\n"
        f"- 逐行对比: ✓ 全部匹配\n"
        f"- 导出的文件名: {download.suggested_filename}\n"
        f"- 文件路径: {temp_path}",
        f"{sheet_name} - 验证结果",
        allure.attachment_type.TEXT,
    )

    logger.info(f"[OK] {sheet_name} 验证通过 - {len(excel_rows)} 行数据全部匹配")